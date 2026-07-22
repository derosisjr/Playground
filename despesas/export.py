#!/usr/bin/env python3
"""
Export da base de Despesas → JSON (agregado), XLSX e CSV
========================================================

Lê `despesas/despesas.sqlite` e gera:
  - despesas-index.json (raiz)   índice AGREGADO consumido pelo painel despesas.html
                                 (séries mensais, top favorecidos, por função/fonte,
                                  alertas fiscais por regras) — único artefato versionado.
  - despesas/despesas.csv        pagamentos completos (análise/Excel; gitignored)
  - despesas/despesas.xlsx       idem, planilha p/ gabinete (gitignored)

O JSON é mantido enxuto: o painel trabalha com agregados; o detalhe linha-a-linha
fica no XLSX/CSV.

Uso:
    python despesas/export.py
"""

import csv
import hashlib
import json
import os
import re
import sqlite3
import sys
from collections import defaultdict
from datetime import datetime

from formato import compacto as _brl_compacto, eh_ente_publico, sem_acento  # camada comum do módulo

for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8")
    except (AttributeError, ValueError):
        pass

AQUI = os.path.dirname(os.path.abspath(__file__))
RAIZ = os.path.dirname(AQUI)
DB_PATH = os.path.join(AQUI, "despesas.sqlite")
CSV_PATH = os.path.join(AQUI, "despesas.csv")
XLSX_PATH = os.path.join(AQUI, "despesas.xlsx")
DADOS_DIR = os.path.join(AQUI, "dados")           # detalhe mensal versionado (consumido pelo painel)
JSON_PATH = os.path.join(RAIZ, "despesas-index.json")
FAV_DIR = os.path.join(RAIZ, "favorecidos")       # raio-X por favorecido (versionado; favorecido.html)
ARVORE_PATH = os.path.join(AQUI, "arvore.json")   # hierarquia função→subfunção→elemento (treemap)

COLUNAS = [
    "ano", "mes", "unidade_gestora", "data", "especie", "empenho", "liquidacao",
    "pagamento", "tipo_pagamento", "elemento_despesa", "subtitulo", "funcao",
    "subfuncao", "programa", "fonte_recurso", "grupo_despesa",
    "documento_favorecido", "nome_favorecido", "valor",
]

# ── Limiares dos alertas fiscais (ajuste fino aqui) ───────────────────────────
TOP_FAVORECIDOS = 300           # quantos favorecidos exportar no JSON
ALERTA_FAV_VALOR = 5_000_000    # favorecido acima disso no acumulado → alerta
ALERTA_FAV_MESES = 6            # ...presente em ao menos N meses (recorrência)
ALERTA_CONCENTRACAO_PCT = 40    # favorecido que domina >X% de uma função → alerta
ALERTA_CONCENTRACAO_MIN = 1_000_000   # ...e move ao menos isso (evita ruído)
ALERTA_PICO_K = 2.5             # mês cujo gasto > média + k·desvio → alerta
ALERTA_EXTRA_VALOR = 1_000_000  # pagamento extra-orçamentário acima disso → alerta

# Regras inteligentes adicionais (calibradas sobre a base do mandato)
# Lei 14.133/2021 art. 75 — limite de dispensa (compras/serviços). MANTER atualizado por decreto.
LIMITE_DISPENSA = 59_906        # usado p/ detectar fracionamento
FRAC_MIN_EMPENHOS = 4           # nº mínimo de empenhos < limite no mesmo favorecido+elemento+ano
FRAC_FATOR_TOTAL = 1.5          # ...somando ao menos FATOR×limite
ALERTA_NOVO_MESES = 6           # "novo" = 1ª aparição nos últimos N meses
ALERTA_NOVO_VALOR = 1_000_000   # ...e já acumulou ao menos isso
ALERTA_YOY_FATOR = 3.0          # crescimento anualizado do ano corrente ≥ FATOR× ano anterior
ALERTA_YOY_VALOR = 1_000_000    # ...e move ao menos isso no ano corrente
ALERTA_PF_VALOR = 100_000       # pessoa física (CPF) em elemento sensível acima disso
CAP_POR_REGRA = 15             # teto de alertas por regra nova (não afogar o painel/e-mail)
MAX_ALERTAS = 100              # era 60

# entes internos/governamentais — excluídos de "novo"/"YoY" (não são fornecedores
# de mercado). Heurística única em formato.eh_ente_publico (compartilhada c/ briefing).
# CPF mascarado: sem "/" e com "*"; elementos sensíveis p/ pessoa física
SQL_EH_CPF = "documento_favorecido NOT LIKE '%/%' AND documento_favorecido LIKE '%*%'"
ELEM_SENSIVEIS = ("%LOCA%IM%", "%TERCEIRIZA%", "%INDENIZA%", "%CONSULTORIA%", "%TERCEIRO%F_SICA%")


def conectar() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def _q(conn, sql, *params):
    return conn.execute(sql, params).fetchall()


# ── Agregações para o painel ──────────────────────────────────────────────────
def agregados(conn) -> dict:
    total = _q(conn, "SELECT COALESCE(SUM(valor),0) s, COUNT(*) n FROM pagamentos")[0]
    anos = _q(conn, "SELECT ano, SUM(valor) s, COUNT(*) n FROM pagamentos GROUP BY ano ORDER BY ano")
    minmax = _q(conn, "SELECT MIN(ano*100+mes) a, MAX(ano*100+mes) b FROM pagamentos")[0]
    dados_ate = _q(conn, "SELECT MAX(data) m FROM pagamentos WHERE data <> ''")[0]["m"]

    def per(a):
        a = int(a); return f"{a//100}-{a%100:02d}"

    series = _q(conn,
        "SELECT ano, mes, SUM(valor) s FROM pagamentos GROUP BY ano, mes ORDER BY ano, mes")
    por_funcao = _q(conn,
        "SELECT COALESCE(NULLIF(funcao,''),'(sem função)') k, SUM(valor) s, COUNT(*) n "
        "FROM pagamentos GROUP BY k ORDER BY s DESC")
    por_elemento = _q(conn,
        "SELECT COALESCE(NULLIF(elemento_despesa,''),'(sem elemento)') k, SUM(valor) s, COUNT(*) n "
        "FROM pagamentos GROUP BY k ORDER BY s DESC LIMIT 50")
    por_fonte = _q(conn,
        "SELECT COALESCE(NULLIF(fonte_recurso,''),'(sem fonte)') k, SUM(valor) s "
        "FROM pagamentos GROUP BY k ORDER BY s DESC LIMIT 50")
    por_unidade = _q(conn,
        "SELECT unidade_gestora k, SUM(valor) s FROM pagamentos GROUP BY k ORDER BY s DESC")
    favorecidos = _q(conn,
        "SELECT nome_favorecido k, documento_favorecido doc, SUM(valor) s, COUNT(*) n, "
        "COUNT(DISTINCT ano*100+mes) meses FROM pagamentos "
        "GROUP BY nome_favorecido, documento_favorecido ORDER BY s DESC LIMIT ?",
        TOP_FAVORECIDOS)

    n_fav = _q(conn, "SELECT COUNT(DISTINCT nome_favorecido) n FROM pagamentos")[0]["n"]

    # serviço da dívida (consumido pelo painel de Endividamento): grupos
    # 32 = juros e encargos, 46 = amortização/refinanciamento (rótulos variam;
    # o código no prefixo do grupo é estável). Credores agrupados por documento
    # (o nome tem grafias divergentes: acento, espaço final).
    FILTRO_DIVIDA = ("(grupo_despesa LIKE '32%' OR grupo_despesa LIKE '46%')")
    divida_anos = _q(conn,
        f"SELECT ano, SUBSTR(grupo_despesa,1,2) g, SUM(valor) s FROM pagamentos "
        f"WHERE {FILTRO_DIVIDA} GROUP BY ano, g ORDER BY ano")
    divida_credores = _q(conn,
        f"SELECT COALESCE(NULLIF(documento_favorecido,''), nome_favorecido) doc, "
        f"MAX(TRIM(nome_favorecido)) k, SUM(valor) s, COUNT(*) n FROM pagamentos "
        f"WHERE {FILTRO_DIVIDA} AND nome_favorecido NOT LIKE '%IPAM%' "
        f"GROUP BY doc ORDER BY s DESC LIMIT 10")

    # série mensal desagregada por função (consumida pelo painel de indicadores)
    serie_funcao = _q(conn,
        "SELECT ano, mes, COALESCE(NULLIF(funcao,''),'(sem função)') f, SUM(valor) s "
        "FROM pagamentos GROUP BY ano, mes, f ORDER BY ano, mes, s DESC")

    return {
        "atualizado_em": datetime.now().isoformat(timespec="seconds"),
        "dados_ate": dados_ate,   # data do pagamento mais recente (aviso de atualidade)
        "periodo": {"de": per(minmax["a"]) if minmax["a"] else None,
                    "ate": per(minmax["b"]) if minmax["b"] else None},
        "totais": {
            "geral": round(total["s"], 2),
            "pagamentos": total["n"],
            "favorecidos": n_fav,
            "por_ano": {str(r["ano"]): round(r["s"], 2) for r in anos},
        },
        "series_mensais": [{"ano": r["ano"], "mes": r["mes"], "valor": round(r["s"], 2)} for r in series],
        "series_mensais_por_funcao": [
            {"ano": r["ano"], "mes": r["mes"], "funcao": r["f"], "valor": round(r["s"], 2)}
            for r in serie_funcao
        ],
        "por_funcao": [{"funcao": r["k"], "valor": round(r["s"], 2), "qtd": r["n"]} for r in por_funcao],
        "por_elemento": [{"elemento": r["k"], "valor": round(r["s"], 2), "qtd": r["n"]} for r in por_elemento],
        "por_fonte": [{"fonte": r["k"], "valor": round(r["s"], 2)} for r in por_fonte],
        "por_unidade": [{"unidade": r["k"], "valor": round(r["s"], 2)} for r in por_unidade],
        "top_favorecidos": [
            {"nome": r["k"], "documento": r["doc"], "valor": round(r["s"], 2),
             "qtd": r["n"], "meses": r["meses"]} for r in favorecidos
        ],
        "servico_divida": {
            "por_ano": [
                {"ano": r["ano"],
                 "tipo": "juros" if r["g"] == "32" else "amortizacao",
                 "valor": round(r["s"], 2)} for r in divida_anos
            ],
            "credores": [
                {"nome": r["k"], "valor": round(r["s"], 2), "qtd": r["n"]}
                for r in divida_credores
            ],
        },
    }


# ── Execução agregada (empenhado → liquidado → pago) ──────────────────────────
# Séries e taxas com os TRÊS estágios que o crawler coleta. As anulações vêm da
# API com valor NEGATIVO (espécie "Anulação"), então SUM(valor) já devolve o
# empenhado/liquidado líquido — sem CASE por espécie. "Restos a pagar" =
# pagamento cujo nº de empenho não existe na base (exercício anterior);
# "extra-orçamentário" = pagamento sem nº de empenho (mesma regra do detalhe).
def execucao_agregada(conn) -> dict:
    serie: dict[int, dict] = {}

    def acumular(tabela, campo):
        for r in _q(conn, f"SELECT ano, mes, ROUND(SUM(valor),2) s FROM {tabela} GROUP BY ano, mes"):
            p = r["ano"] * 100 + r["mes"]
            serie.setdefault(p, {"ano": r["ano"], "mes": r["mes"]})[campo] = r["s"]

    acumular("empenhos", "empenhado")
    acumular("liquidacoes", "liquidado")
    acumular("pagamentos", "pago")

    # restos a pagar e extra-orçamentário dentro do pago de cada mês
    for r in _q(conn, """
        SELECT pg.ano, pg.mes,
               ROUND(SUM(CASE WHEN pg.empenho IS NULL OR pg.empenho = '' THEN pg.valor ELSE 0 END),2) extra,
               ROUND(SUM(CASE WHEN pg.empenho IS NOT NULL AND pg.empenho <> '' AND e.empenho IS NULL
                             THEN pg.valor ELSE 0 END),2) restos
        FROM pagamentos pg
        LEFT JOIN (SELECT DISTINCT unidade_gestora, empenho FROM empenhos) e
          ON e.unidade_gestora = pg.unidade_gestora AND e.empenho = pg.empenho
        GROUP BY pg.ano, pg.mes"""):
        p = r["ano"] * 100 + r["mes"]
        if p in serie:
            serie[p]["restos"] = r["restos"]
            serie[p]["extra"] = r["extra"]

    linhas = [serie[p] for p in sorted(serie)]
    for l in linhas:
        for c in ("empenhado", "liquidado", "pago", "restos", "extra"):
            l.setdefault(c, 0)

    # totais por ano (movimento de caixa/competência de cada estágio)
    por_ano = {}
    for l in linhas:
        a = por_ano.setdefault(str(l["ano"]), {"empenhado": 0, "liquidado": 0, "pago": 0,
                                               "restos": 0, "extra": 0})
        for c in ("empenhado", "liquidado", "pago", "restos", "extra"):
            a[c] = round(a[c] + l[c], 2)

    # Taxa de execução por ano DO EMPENHO: dos empenhos emitidos no ano, quanto já
    # foi liquidado/pago (casando pela chave do empenho, como no detalhe). Só é
    # publicada quando a base de empenhos do ano está íntegra: empenhos globais
    # (folha, previdência) emitidos ANTES do início da base (dez do exercício
    # anterior) entram só como anulação/reforço e subcontam o empenhado — nesses
    # anos o pago casado excede o empenhado e a taxa sairia >100% (enganosa).
    for r in _q(conn, """
        SELECT e.ano, ROUND(SUM(e.empenhado),2) emp,
               ROUND(SUM(COALESCE(l.liquidado,0)),2) liq,
               ROUND(SUM(COALESCE(p.pago,0)),2) pago
        FROM (SELECT unidade_gestora, empenho, MIN(ano) ano, ROUND(SUM(valor),2) empenhado
              FROM empenhos GROUP BY unidade_gestora, empenho) e
        LEFT JOIN (SELECT unidade_gestora, empenho, ROUND(SUM(valor),2) liquidado
                   FROM liquidacoes GROUP BY unidade_gestora, empenho) l
          ON l.unidade_gestora = e.unidade_gestora AND l.empenho = e.empenho
        LEFT JOIN (SELECT unidade_gestora, empenho, ROUND(SUM(valor),2) pago
                   FROM pagamentos WHERE empenho IS NOT NULL AND empenho <> ''
                   GROUP BY unidade_gestora, empenho) p
          ON p.unidade_gestora = e.unidade_gestora AND p.empenho = e.empenho
        GROUP BY e.ano"""):
        a = por_ano.get(str(r["ano"]))
        if not a or not r["emp"]:
            continue
        if r["pago"] > r["emp"]:
            a["empenho_incompleto"] = True   # base sem os empenhos iniciais do exercício
        else:
            a["taxa_liquidacao"] = round(100 * r["liq"] / r["emp"], 1)
            a["taxa_pagamento"] = round(100 * r["pago"] / r["emp"], 1)

    # tríade por função (top 12 pelo empenhado)
    fn: dict[str, dict] = {}
    for tabela, campo in (("empenhos", "empenhado"), ("liquidacoes", "liquidado"),
                          ("pagamentos", "pago")):
        for r in _q(conn,
            f"SELECT COALESCE(NULLIF(funcao,''),'(sem função)') k, ROUND(SUM(valor),2) s "
            f"FROM {tabela} GROUP BY k"):
            fn.setdefault(r["k"], {"funcao": r["k"]})[campo] = r["s"]
    por_funcao = sorted(fn.values(), key=lambda x: -(x.get("empenhado") or 0))[:12]
    for f in por_funcao:
        for c in ("empenhado", "liquidado", "pago"):
            f.setdefault(c, 0)

    return {"serie": linhas, "por_ano": por_ano, "por_funcao": por_funcao}


# ── Resumo narrativo ("Em resumo") ────────────────────────────────────────────
# Texto determinístico (sem IA) exibido no topo do painel: último mês completo vs
# média móvel, função que mais variou, acumulado do ano vs mesmo período anterior
# e alertas ativos. Frases montadas por template — auditável e sem custo de token.
MESES_NOME = ["janeiro", "fevereiro", "março", "abril", "maio", "junho", "julho",
              "agosto", "setembro", "outubro", "novembro", "dezembro"]
MESES_ABREV = ["jan", "fev", "mar", "abr", "mai", "jun", "jul", "ago", "set", "out", "nov", "dez"]


def resumo_narrativo(conn, indice: dict) -> dict | None:
    series = indice.get("series_mensais") or []
    if len(series) < 3:
        return None

    # último mês COMPLETO: descarta o mês corrente (parcial) se for o último da série
    hoje = datetime.now()
    idx = len(series) - 1
    if series[idx]["ano"] == hoje.year and series[idx]["mes"] == hoje.month:
        idx -= 1
    ref = series[idx]

    # vs média dos até 12 meses anteriores
    anteriores = series[max(0, idx - 12):idx]
    media = sum(s["valor"] for s in anteriores) / len(anteriores) if anteriores else None
    delta_media_pct = round((ref["valor"] - media) / media * 100) if media else None

    # função com maior variação absoluta vs o mês anterior
    destaque = None
    if idx >= 1:
        ant = series[idx - 1]
        rows = _q(conn, """
            SELECT COALESCE(NULLIF(funcao,''),'(sem função)') k,
                   SUM(CASE WHEN ano=? AND mes=? THEN valor ELSE 0 END) atual,
                   SUM(CASE WHEN ano=? AND mes=? THEN valor ELSE 0 END) anterior
            FROM pagamentos
            WHERE (ano=? AND mes=?) OR (ano=? AND mes=?)
            GROUP BY k""",
            ref["ano"], ref["mes"], ant["ano"], ant["mes"],
            ref["ano"], ref["mes"], ant["ano"], ant["mes"])
        difs = [(r["k"], (r["atual"] or 0) - (r["anterior"] or 0)) for r in rows]
        if difs:
            k, d = max(difs, key=lambda x: abs(x[1]))
            if abs(d) >= 1_000_000:  # só destaca variação relevante
                nome = re.sub(r"^\d+\s*-\s*", "", k).strip().capitalize()
                destaque = {"funcao": nome, "delta": round(d, 2)}

    # acumulado do ano (jan..mês de referência) vs mesmo período do ano anterior
    yoy = None
    atual = sum(s["valor"] for s in series if s["ano"] == ref["ano"] and s["mes"] <= ref["mes"])
    passado = sum(s["valor"] for s in series if s["ano"] == ref["ano"] - 1 and s["mes"] <= ref["mes"])
    if passado:
        yoy = {"atual": round(atual, 2), "anterior": round(passado, 2),
               "pct": round((atual - passado) / passado * 100)}

    ativos = sum(1 for a in indice.get("alertas", [])
                 if a.get("severidade") in ("alta", "media"))

    frases = []
    f1 = f"Em {MESES_NOME[ref['mes'] - 1]}, a Prefeitura pagou {_brl_compacto(ref['valor'])}"
    if delta_media_pct is not None and delta_media_pct != 0:
        f1 += (f" — {abs(delta_media_pct)}% "
               f"{'acima' if delta_media_pct > 0 else 'abaixo'} da média dos 12 meses anteriores")
    frases.append(f1 + ".")
    if destaque:
        sobe = destaque["delta"] > 0
        frases.append(f"{destaque['funcao']} teve a maior variação: "
                      f"{'+' if sobe else '−'}{_brl_compacto(abs(destaque['delta']))} "
                      f"em relação ao mês anterior.")
    if yoy:
        frases.append(f"No acumulado de {ref['ano']} (jan–{MESES_ABREV[ref['mes'] - 1]}), "
                      f"o pago soma {_brl_compacto(yoy['atual'])}, "
                      f"{abs(yoy['pct'])}% {'acima' if yoy['pct'] >= 0 else 'abaixo'} "
                      f"do mesmo período de {ref['ano'] - 1}.")
    exe = (indice.get("execucao") or {}).get("por_ano", {}).get(str(ref["ano"]))
    if exe:
        partes = []
        if exe.get("taxa_pagamento") is not None:
            partes.append(f"Do empenhado em {ref['ano']}, {exe['taxa_liquidacao']:.0f}% já foi "
                          f"liquidado e {exe['taxa_pagamento']:.0f}% pago")
        if exe.get("restos"):
            partes.append(f"{_brl_compacto(exe['restos'])} quitaram restos a pagar "
                          f"de exercícios anteriores")
        if partes:
            frases.append("; ".join(partes) + ".")
    if ativos:
        frases.append(f"Há {ativos} alertas fiscais ativos de severidade alta ou média.")

    return {
        "texto": " ".join(frases),
        "mes_ref": {"ano": ref["ano"], "mes": ref["mes"], "valor": ref["valor"]},
        "delta_media_pct": delta_media_pct,
        "funcao_destaque": destaque,
        "yoy": yoy,
        "alertas_ativos": ativos,
    }


# ── Detalhe por ano (retrospectiva.html) ──────────────────────────────────────
def anos_detalhe(conn) -> dict:
    """Top funções e favorecidos POR ANO — alimenta a página-história do ano."""
    out = {}
    for r in _q(conn, "SELECT DISTINCT ano FROM pagamentos ORDER BY ano"):
        a = r["ano"]
        fn = [{"funcao": x["k"], "valor": round(x["s"], 2)}
              for x in _q(conn,
                  "SELECT COALESCE(NULLIF(funcao,''),'(sem função)') k, SUM(valor) s "
                  "FROM pagamentos WHERE ano=? GROUP BY k ORDER BY s DESC LIMIT 8", a)]
        fav = [{"nome": x["k"], "documento": x["doc"], "valor": round(x["s"], 2)}
               for x in _q(conn,
                   "SELECT nome_favorecido k, documento_favorecido doc, SUM(valor) s "
                   "FROM pagamentos WHERE ano=? GROUP BY k, doc ORDER BY s DESC LIMIT 6", a)]
        out[str(a)] = {"por_funcao": fn, "top_favorecidos": fav}
    return out


# ── Raio-X por favorecido (favorecidos/<slug>.json) ───────────────────────────
# Um dossiê estático por favorecido do top-N: série mensal própria, por função,
# últimos pagamentos e alertas — consumido por favorecido.html?f=<slug>.
def _slug_fav(nome: str, doc: str) -> str:
    """CNPJ completo vira os 14 dígitos; CPF mascarado/sem doc vira hash curto do par."""
    dig = re.sub(r"\D", "", doc or "")
    if dig and "*" not in (doc or ""):
        return dig
    return hashlib.md5(f"{nome}|{doc}".encode("utf-8")).hexdigest()[:12]


def exportar_favorecidos(conn, indice: dict) -> int:
    os.makedirs(FAV_DIR, exist_ok=True)
    conn.execute("CREATE INDEX IF NOT EXISTS ix_pagamentos_fav "
                 "ON pagamentos(nome_favorecido, documento_favorecido)")
    tops = indice.get("top_favorecidos") or []
    validos = set()
    for rank, f in enumerate(tops, 1):
        nome, doc = f["nome"], f.get("documento") or ""
        slug = _slug_fav(nome, doc)
        f["slug"] = slug           # o painel/paleta usam p/ linkar o raio-X
        validos.add(slug + ".json")

        serie = [{"ano": r["ano"], "mes": r["mes"], "valor": round(r["s"], 2)}
                 for r in _q(conn,
                     "SELECT ano, mes, SUM(valor) s FROM pagamentos "
                     "WHERE nome_favorecido=? AND documento_favorecido=? "
                     "GROUP BY ano, mes ORDER BY ano, mes", nome, doc)]
        por_ano = {}
        for s in serie:
            por_ano[str(s["ano"])] = round(por_ano.get(str(s["ano"]), 0) + s["valor"], 2)
        funcoes = [{"funcao": r["k"], "valor": round(r["s"], 2)}
                   for r in _q(conn,
                       "SELECT COALESCE(NULLIF(funcao,''),'(sem função)') k, SUM(valor) s "
                       "FROM pagamentos WHERE nome_favorecido=? AND documento_favorecido=? "
                       "GROUP BY k ORDER BY s DESC LIMIT 6", nome, doc)]
        ultimos = [{"data": r["data"], "valor": round(r["valor"], 2), "funcao": r["funcao"],
                    "elemento": r["elemento_despesa"], "unidade": r["unidade_gestora"]}
                   for r in _q(conn,
                       "SELECT data, valor, funcao, elemento_despesa, unidade_gestora "
                       "FROM pagamentos WHERE nome_favorecido=? AND documento_favorecido=? "
                       "ORDER BY data DESC, valor DESC LIMIT 50", nome, doc)]
        meus_alertas = [a for a in indice.get("alertas", [])
                        if (a.get("filtro") or {}).get("favorecido") == nome]

        with open(os.path.join(FAV_DIR, slug + ".json"), "w", encoding="utf-8") as fp:
            json.dump({
                "nome": nome, "documento": doc, "slug": slug, "rank": rank,
                "total": f["valor"], "qtd": f["qtd"], "meses": f["meses"],
                "por_ano": por_ano, "serie_mensal": serie, "por_funcao": funcoes,
                "ultimos_pagamentos": ultimos, "alertas": meus_alertas,
                "atualizado_em": indice.get("atualizado_em"),
            }, fp, ensure_ascii=False, separators=(",", ":"))

    # remove dossiês de quem saiu do top-N
    for nome_arq in os.listdir(FAV_DIR):
        if nome_arq.endswith(".json") and nome_arq not in validos:
            os.remove(os.path.join(FAV_DIR, nome_arq))
    return len(validos)


# ── Alertas fiscais por regras ────────────────────────────────────────────────
def alertas(conn) -> list[dict]:
    out = []

    # 1) Favorecido recorrente de alto valor
    for r in _q(conn,
        "SELECT nome_favorecido k, documento_favorecido doc, SUM(valor) s, "
        "COUNT(DISTINCT ano*100+mes) meses FROM pagamentos "
        "GROUP BY nome_favorecido, documento_favorecido "
        "HAVING s >= ? AND meses >= ? ORDER BY s DESC LIMIT ?",
        ALERTA_FAV_VALOR, ALERTA_FAV_MESES, CAP_POR_REGRA):
        out.append({
            "tipo": "favorecido_recorrente",
            "severidade": "alta",
            "titulo": f"Favorecido recorrente de alto valor: {r['k']}",
            "detalhe": f"R$ {r['s']:,.2f} acumulados em {r['meses']} meses.",
            "valor": round(r["s"], 2),
            "filtro": {"favorecido": r["k"]},
        })

    # 2) Concentração: favorecido que domina uma função
    tot_funcao = {r["k"]: r["s"] for r in _q(conn,
        "SELECT COALESCE(NULLIF(funcao,''),'(sem função)') k, SUM(valor) s "
        "FROM pagamentos GROUP BY k")}
    n_conc = 0
    for r in _q(conn,
        "SELECT COALESCE(NULLIF(funcao,''),'(sem função)') f, nome_favorecido k, SUM(valor) s "
        "FROM pagamentos GROUP BY f, nome_favorecido HAVING s >= ? ORDER BY s DESC",
        ALERTA_CONCENTRACAO_MIN):
        base = tot_funcao.get(r["f"], 0) or 1
        pct = 100 * r["s"] / base
        if pct >= ALERTA_CONCENTRACAO_PCT and r["f"] != "(sem função)":
            if n_conc >= CAP_POR_REGRA:
                break
            n_conc += 1
            out.append({
                "tipo": "concentracao",
                "severidade": "media",
                "titulo": f"Concentração: {r['k']} concentra {pct:.0f}% da função {r['f']}",
                "detalhe": f"R$ {r['s']:,.2f} de R$ {base:,.2f} na função.",
                "valor": round(r["s"], 2),
                "filtro": {"favorecido": r["k"], "funcao": r["f"]},
            })

    # 3) Pico mensal (sobre o total geral por mês)
    serie = _q(conn, "SELECT ano, mes, SUM(valor) s FROM pagamentos GROUP BY ano, mes ORDER BY ano, mes")
    vals = [r["s"] for r in serie]
    if len(vals) >= 4:
        media = sum(vals) / len(vals)
        desvio = (sum((v - media) ** 2 for v in vals) / len(vals)) ** 0.5
        limite = media + ALERTA_PICO_K * desvio
        for r in serie:
            if r["s"] > limite:
                out.append({
                    "tipo": "pico_mensal",
                    "severidade": "media",
                    "titulo": f"Pico de gasto em {r['ano']}-{r['mes']:02d}",
                    "detalhe": f"R$ {r['s']:,.2f} (média mensal R$ {media:,.2f}).",
                    "valor": round(r["s"], 2),
                    "filtro": {"ano": r["ano"], "mes": r["mes"]},
                })

    # 4) Extra-orçamentário relevante
    for r in _q(conn,
        "SELECT nome_favorecido k, elemento_despesa e, SUM(valor) s, COUNT(*) n "
        "FROM pagamentos WHERE tipo_pagamento LIKE '%Extra%' "
        "GROUP BY nome_favorecido, elemento_despesa HAVING s >= ? ORDER BY s DESC LIMIT 20",
        ALERTA_EXTRA_VALOR):
        out.append({
            "tipo": "extra_orcamentario",
            "severidade": "baixa",
            "titulo": f"Extra-orçamentário relevante: {r['k']}",
            "detalhe": f"R$ {r['s']:,.2f} em {r['n']} pagamentos — {(r['e'] or '')[:60]}.",
            "valor": round(r["s"], 2),
            "filtro": {"favorecido": r["k"]},
        })

    # 5) Possível fracionamento de despesa (vários empenhos logo abaixo do limite de dispensa)
    n = 0
    for r in _q(conn,
        "SELECT nome_favorecido k, elemento_despesa e, ano, COUNT(*) qt, SUM(valor) s "
        "FROM empenhos WHERE valor < ? AND valor > ? "
        "GROUP BY nome_favorecido, elemento_despesa, ano "
        "HAVING qt >= ? AND s >= ? ORDER BY s DESC",
        LIMITE_DISPENSA, LIMITE_DISPENSA * 0.4, FRAC_MIN_EMPENHOS, LIMITE_DISPENSA * FRAC_FATOR_TOTAL):
        if n >= CAP_POR_REGRA:
            break
        n += 1
        out.append({
            "tipo": "fracionamento",
            "severidade": "alta",
            "titulo": f"Possível fracionamento: {r['qt']} empenhos a {r['k']}",
            "detalhe": (f"R$ {r['s']:,.2f} em {r['qt']} empenhos abaixo de R$ {LIMITE_DISPENSA:,.2f} "
                        f"({r['ano']}) — {(r['e'] or '')[:50]}."),
            "valor": round(r["s"], 2),
            "filtro": {"favorecido": r["k"], "elemento": r["e"], "tipo_doc": "pj"},
        })

    # 6) Favorecido novo (1ª aparição recente) de alto valor
    ultimo = _q(conn, "SELECT MAX(ano*100+mes) m FROM pagamentos")[0]["m"]
    if ultimo:
        ay, am = ultimo // 100, ultimo % 100
        cm = am - ALERTA_NOVO_MESES
        cy = ay
        while cm <= 0:
            cm += 12
            cy -= 1
        corte = cy * 100 + cm
        n = 0
        for r in _q(conn,
            "SELECT nome_favorecido k, documento_favorecido doc, SUM(valor) s, "
            "MIN(ano*100+mes) ini FROM pagamentos GROUP BY nome_favorecido, documento_favorecido "
            "HAVING ini >= ? AND s >= ? ORDER BY s DESC",
            corte, ALERTA_NOVO_VALOR):
            if eh_ente_publico(r["k"]):
                continue
            if n >= CAP_POR_REGRA:
                break
            n += 1
            out.append({
                "tipo": "favorecido_novo",
                "severidade": "media",
                "titulo": f"Favorecido novo de alto valor: {r['k']}",
                "detalhe": (f"R$ {r['s']:,.2f} desde {r['ini'] // 100}-{r['ini'] % 100:02d} "
                            f"(1ª aparição na base)."),
                "valor": round(r["s"], 2),
                "filtro": {"favorecido": r["k"]},
            })

    # 7) Crescimento anômalo ano-a-ano (anualizando o ano corrente)
    anos = [r["ano"] for r in _q(conn, "SELECT DISTINCT ano FROM pagamentos ORDER BY ano")]
    if len(anos) >= 2:
        ant, atual = anos[-2], anos[-1]
        meses_dec = _q(conn, "SELECT MAX(mes) m FROM pagamentos WHERE ano = ?", atual)[0]["m"] or 12
        fator_anual = 12.0 / meses_dec
        n = 0
        for r in _q(conn,
            "SELECT nome_favorecido k, "
            "SUM(CASE WHEN ano=? THEN valor ELSE 0 END) v_ant, "
            "SUM(CASE WHEN ano=? THEN valor ELSE 0 END) v_atual "
            "FROM pagamentos GROUP BY nome_favorecido "
            "HAVING v_ant > 0 AND v_atual >= ? ORDER BY v_atual DESC",
            ant, atual, ALERTA_YOY_VALOR):
            if eh_ente_publico(r["k"]):
                continue
            proj = r["v_atual"] * fator_anual
            if proj < ALERTA_YOY_FATOR * r["v_ant"]:
                continue
            if n >= CAP_POR_REGRA:
                break
            n += 1
            out.append({
                "tipo": "crescimento_yoy",
                "severidade": "media",
                "titulo": f"Crescimento anômalo: {r['k']}",
                "detalhe": (f"R$ {r['v_atual']:,.2f} em {atual} (proj. anual R$ {proj:,.2f}) "
                            f"vs R$ {r['v_ant']:,.2f} em {ant} — {proj / r['v_ant']:.1f}×."),
                "valor": round(r["v_atual"], 2),
                "filtro": {"favorecido": r["k"]},
            })

    # 8) Pessoa física (CPF) recebendo em elemento sensível
    cond_elem = " OR ".join(f"elemento_despesa LIKE '{p}'" for p in ELEM_SENSIVEIS)
    n = 0
    for r in _q(conn,
        "SELECT nome_favorecido k, elemento_despesa e, SUM(valor) s, COUNT(*) qt "
        f"FROM pagamentos WHERE ({SQL_EH_CPF}) AND ({cond_elem}) "
        "GROUP BY nome_favorecido, elemento_despesa HAVING s >= ? ORDER BY s DESC",
        ALERTA_PF_VALOR):
        if n >= CAP_POR_REGRA:
            break
        n += 1
        elem_curto = (r["e"] or "").split(" - ", 1)[-1][:40]
        out.append({
            "tipo": "pf_sensivel",
            "severidade": "alta",
            "titulo": f"Pessoa física em {elem_curto}: {r['k']}",
            "detalhe": f"R$ {r['s']:,.2f} em {r['qt']} pagamentos a pessoa física — {(r['e'] or '')[:50]}.",
            "valor": round(r["s"], 2),
            "filtro": {"favorecido": r["k"], "elemento": r["e"], "tipo_doc": "pf"},
        })

    ordem = {"alta": 0, "media": 1, "baixa": 2}
    out.sort(key=lambda a: (ordem.get(a["severidade"], 9), -a["valor"]))
    return out[:MAX_ALERTAS]


# ── Árvore do gasto (função → subfunção → elemento) p/ o treemap ──────────────
# Gravada em arquivo próprio (despesas/arvore.json) para não inchar o índice:
# são ~100 subfunções; os elementos são limitados ao top-N + "Demais" por
# subfunção. Folhas com soma ≤ 0 (anulações líquidas) são absorvidas em "Demais".
TOP_ELEM_ARVORE = 10

def exportar_arvore(conn, indice: dict) -> int:
    rows = _q(conn, """
        SELECT COALESCE(NULLIF(funcao,''),'(sem função)') f,
               COALESCE(NULLIF(subfuncao,''),'(sem subfunção)') sf,
               COALESCE(NULLIF(elemento_despesa,''),'(sem elemento)') e,
               ROUND(SUM(valor),2) s
        FROM pagamentos GROUP BY f, sf, e""")
    arvore: dict[str, dict] = {}
    for r in rows:
        fn = arvore.setdefault(r["f"], {"n": r["f"], "v": 0, "f": {}})
        sf = fn["f"].setdefault(r["sf"], {"n": r["sf"], "v": 0, "f": []})
        fn["v"] = round(fn["v"] + r["s"], 2)
        sf["v"] = round(sf["v"] + r["s"], 2)
        sf["f"].append({"n": r["e"], "v": r["s"]})

    saida = []
    for fn in sorted(arvore.values(), key=lambda x: -x["v"]):
        subs = []
        for sf in sorted(fn["f"].values(), key=lambda x: -x["v"]):
            folhas = sorted(sf["f"], key=lambda x: -x["v"])
            top = [x for x in folhas[:TOP_ELEM_ARVORE] if x["v"] > 0]
            resto = round(sum(x["v"] for x in folhas[TOP_ELEM_ARVORE:]) +
                          sum(x["v"] for x in folhas[:TOP_ELEM_ARVORE] if x["v"] <= 0), 2)
            if resto > 0:
                top.append({"n": "Demais elementos", "v": resto})
            subs.append({"n": sf["n"], "v": sf["v"], "f": top})
        saida.append({"n": fn["n"], "v": fn["v"], "f": subs})

    with open(ARVORE_PATH, "w", encoding="utf-8") as fp:
        json.dump({"atualizado_em": indice.get("atualizado_em"),
                   "total": indice.get("totais", {}).get("geral"),
                   "arvore": saida}, fp, ensure_ascii=False, separators=(",", ":"))
    return len(saida)


# ── Execução por empenho (empenhado / liquidado / pago) ───────────────────────
# Detalhe que o painel carrega sob demanda. Unidade da linha = o EMPENHO; junta os
# três estágios (empenhos+liquidacoes+pagamentos) por (unidade_gestora, empenho).
# Inclui também o que não casa com empenho da base, para não esconder nada:
#   - "Restos a pagar": pagamento cujo empenho é de exercício anterior (fora da base)
#   - "Extra-orçamentário": pagamento sem empenho (retenções/tributos)
CAMPOS_DETALHE = [
    "data", "unidade_gestora", "tipo", "nome_favorecido", "documento_favorecido",
    "funcao", "subfuncao", "programa", "elemento_despesa", "fonte_recurso",
    "grupo_despesa", "empenho", "empenhado", "liquidado", "pago",
]
COLS_NUM = {"empenhado", "liquidado", "pago"}


def montar_execucao(conn) -> list[dict]:
    """Devolve uma linha por empenho (+ restos/extra-orçamentário), com _periodo (AAAAMM)."""
    rows = []
    # 1) Empenhos da base, com liquidado e pago somados por empenho.
    for r in _q(conn, """
        SELECT e.unidade_gestora, e.empenho, e.periodo, e.data, e.favorecido, e.doc,
               e.funcao, e.subfuncao, e.programa, e.elemento, e.fonte, e.grupo,
               e.empenhado,
               COALESCE(l.liquidado, 0) liquidado, COALESCE(p.pago, 0) pago
        FROM (
            SELECT unidade_gestora, empenho, MIN(ano*100+mes) periodo, MIN(data) data,
                   MAX(nome_favorecido) favorecido, MAX(documento_favorecido) doc,
                   MAX(funcao) funcao, MAX(subfuncao) subfuncao, MAX(programa) programa,
                   MAX(elemento_despesa) elemento, MAX(fonte_recurso) fonte,
                   MAX(grupo_despesa) grupo, ROUND(SUM(valor), 2) empenhado
            FROM empenhos GROUP BY unidade_gestora, empenho
        ) e
        LEFT JOIN (SELECT unidade_gestora, empenho, ROUND(SUM(valor),2) liquidado
                   FROM liquidacoes GROUP BY unidade_gestora, empenho) l
          ON l.unidade_gestora = e.unidade_gestora AND l.empenho = e.empenho
        LEFT JOIN (SELECT unidade_gestora, empenho, ROUND(SUM(valor),2) pago
                   FROM pagamentos WHERE empenho IS NOT NULL AND empenho <> ''
                   GROUP BY unidade_gestora, empenho) p
          ON p.unidade_gestora = e.unidade_gestora AND p.empenho = e.empenho
    """):
        rows.append({"_periodo": r["periodo"], "data": r["data"],
                     "unidade_gestora": r["unidade_gestora"], "tipo": "Empenho",
                     "nome_favorecido": r["favorecido"], "documento_favorecido": r["doc"],
                     "funcao": r["funcao"], "subfuncao": r["subfuncao"], "programa": r["programa"],
                     "elemento_despesa": r["elemento"], "fonte_recurso": r["fonte"],
                     "grupo_despesa": r["grupo"], "empenho": r["empenho"],
                     "empenhado": r["empenhado"], "liquidado": r["liquidado"], "pago": r["pago"]})

    # 2) Pagamentos que NÃO casam com empenho da base (restos a pagar / extra-orçamentário).
    for r in _q(conn, """
        SELECT pg.unidade_gestora, pg.empenho, (pg.ano*100+pg.mes) periodo, pg.data,
               pg.nome_favorecido, pg.documento_favorecido, pg.funcao, pg.subfuncao,
               pg.programa, pg.elemento_despesa, pg.fonte_recurso, pg.grupo_despesa,
               ROUND(SUM(pg.valor),2) pago,
               (pg.empenho IS NULL OR pg.empenho = '') sem_emp
        FROM pagamentos pg
        LEFT JOIN (SELECT DISTINCT unidade_gestora, empenho FROM empenhos) e
          ON e.unidade_gestora = pg.unidade_gestora AND e.empenho = pg.empenho
        WHERE e.empenho IS NULL
        GROUP BY pg.unidade_gestora, pg.empenho, pg.ano, pg.mes, pg.data,
                 pg.nome_favorecido, pg.elemento_despesa
    """):
        sem_emp = r["sem_emp"]
        rows.append({"_periodo": r["periodo"], "data": r["data"],
                     "unidade_gestora": r["unidade_gestora"],
                     "tipo": "Extra-orçamentário" if sem_emp else "Restos a pagar",
                     "nome_favorecido": r["nome_favorecido"], "documento_favorecido": r["documento_favorecido"],
                     "funcao": r["funcao"], "subfuncao": r["subfuncao"], "programa": r["programa"],
                     "elemento_despesa": r["elemento_despesa"], "fonte_recurso": r["fonte_recurso"],
                     "grupo_despesa": r["grupo_despesa"], "empenho": r["empenho"] or "",
                     "empenhado": None, "liquidado": None, "pago": r["pago"]})

    rows.sort(key=lambda x: (x["_periodo"], x["data"] or "", -(x["pago"] or 0)))
    return rows


# ── Exports tabulares (completos, gitignored) ─────────────────────────────────
def exportar_csv(rows: list[dict]) -> None:
    with open(CSV_PATH, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(CAMPOS_DETALHE)
        for r in rows:
            w.writerow([r.get(c) for c in CAMPOS_DETALHE])


def exportar_xlsx(rows: list[dict]) -> None:
    from openpyxl import Workbook
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
    from openpyxl.utils import get_column_letter

    def limpar(v):
        return ILLEGAL_CHARACTERS_RE.sub("", v) if isinstance(v, str) else v

    wb = Workbook()
    ws = wb.active
    ws.title = "Execução por empenho"
    ws.append(CAMPOS_DETALHE)
    for r in rows:
        ws.append([limpar(r.get(c)) for c in CAMPOS_DETALHE])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(CAMPOS_DETALHE))}{len(rows)+1}"
    larguras = {"unidade_gestora": 32, "elemento_despesa": 40, "nome_favorecido": 36,
                "funcao": 22, "fonte_recurso": 26, "grupo_despesa": 26, "programa": 28,
                "empenhado": 14, "liquidado": 14, "pago": 14, "tipo": 16}
    for i, c in enumerate(CAMPOS_DETALHE, 1):
        ws.column_dimensions[get_column_letter(i)].width = larguras.get(c, 14)
    wb.save(XLSX_PATH)


def exportar_detalhe_mensal(rows: list[dict]) -> list[dict]:
    """Grava despesas/dados/AAAA-MM.json (um por mês de competência) e devolve o manifesto."""
    os.makedirs(DADOS_DIR, exist_ok=True)
    por_mes: dict[int, list] = {}
    for r in rows:
        por_mes.setdefault(r["_periodo"], []).append([r.get(c) for c in CAMPOS_DETALHE])

    validos = {f"{p // 100}-{p % 100:02d}.json" for p in por_mes} | set(INDICES_LEVES)
    for nome in os.listdir(DADOS_DIR):
        if nome.endswith(".json") and nome not in validos:
            os.remove(os.path.join(DADOS_DIR, nome))

    manifesto = []
    for periodo in sorted(por_mes):
        ano, mes = periodo // 100, periodo % 100
        linhas = por_mes[periodo]
        arquivo = f"{ano}-{mes:02d}.json"
        with open(os.path.join(DADOS_DIR, arquivo), "w", encoding="utf-8") as f:
            json.dump({"campos": CAMPOS_DETALHE, "linhas": linhas},
                      f, ensure_ascii=False, separators=(",", ":"))
        soma_pago = round(sum((l[-1] or 0) for l in linhas), 2)
        manifesto.append({"ano": ano, "mes": mes, "n": len(linhas),
                          "valor": soma_pago, "arquivo": f"despesas/dados/{arquivo}"})
    return manifesto


# ── Índices leves (evitam baixar os ~20 MB de detalhe no painel) ──────────────
# Três arquivos pequenos em despesas/dados/:
#   elementos.json          opções do filtro de elemento por tipo de documento
#   pf-resumo.json          agregado das pessoas físicas (modo PF sem detalhe)
#   indice-favorecidos.json favorecido → meses onde aparece (a ficha baixa SÓ esses)
# Chave do favorecido = dígitos do documento; sem documento, nome sem acento em
# minúsculas — TEM de casar com chaveFavorecido() do despesas-app.js.
INDICES_LEVES = ("elementos.json", "pf-resumo.json", "indice-favorecidos.json")
PF_RESUMO_TOP = 1000


def _eh_cpf(doc) -> bool:
    return bool(doc) and "/" not in doc and "*" in doc


def _chave_fav(nome, doc) -> str:
    dig = re.sub(r"\D", "", doc or "")
    return dig if dig else sem_acento(nome).lower()


def exportar_indices_leves(rows: list[dict]) -> None:
    elementos = {"todos": set(), "pf": set(), "pj": set()}
    pf: dict[tuple, dict] = {}
    fav_meses: dict[str, set] = {}
    todos_meses = set()

    for r in rows:
        doc = r.get("documento_favorecido") or ""
        nome = r.get("nome_favorecido") or ""
        elem = r.get("elemento_despesa")
        periodo = r["_periodo"]
        todos_meses.add(periodo)
        if elem:
            elementos["todos"].add(elem)
            if _eh_cpf(doc):
                elementos["pf"].add(elem)
            elif "/" in doc:
                elementos["pj"].add(elem)
        fav_meses.setdefault(_chave_fav(nome, doc), set()).add(periodo)
        if _eh_cpf(doc):
            o = pf.setdefault((nome, doc), {"nome": nome, "documento": doc,
                                            "valor": 0.0, "qtd": 0, "meses": set()})
            o["valor"] += r.get("pago") or 0
            o["qtd"] += 1
            if r.get("data"):
                o["meses"].add(str(r["data"])[:7])

    with open(os.path.join(DADOS_DIR, "elementos.json"), "w", encoding="utf-8") as f:
        json.dump({k: sorted(v) for k, v in elementos.items()},
                  f, ensure_ascii=False, separators=(",", ":"))

    itens_pf = sorted(pf.values(), key=lambda x: -x["valor"])[:PF_RESUMO_TOP]
    with open(os.path.join(DADOS_DIR, "pf-resumo.json"), "w", encoding="utf-8") as f:
        json.dump({"itens": [{"nome": o["nome"], "documento": o["documento"],
                              "valor": round(o["valor"], 2), "qtd": o["qtd"],
                              "meses": len(o["meses"])} for o in itens_pf]},
                  f, ensure_ascii=False, separators=(",", ":"))

    # quem aparece em quase todos os meses fica FORA do índice (baixar tudo equivale);
    # chave ausente no painel = fallback para todos os meses.
    corte = max(1, len(todos_meses) - 2)
    fav = {k: sorted(v) for k, v in fav_meses.items() if len(v) < corte}
    with open(os.path.join(DADOS_DIR, "indice-favorecidos.json"), "w", encoding="utf-8") as f:
        json.dump({"meses_total": len(todos_meses), "fav": fav},
                  f, ensure_ascii=False, separators=(",", ":"))


def main():
    if not os.path.exists(DB_PATH):
        print(f"Banco não encontrado: {DB_PATH}. Rode o crawler primeiro.", file=sys.stderr)
        sys.exit(1)
    conn = conectar()
    indice = agregados(conn)
    indice["execucao"] = execucao_agregada(conn)
    indice["alertas"] = alertas(conn)
    indice["resumo"] = resumo_narrativo(conn, indice)
    indice["anos_detalhe"] = anos_detalhe(conn)
    n_fav_raiox = exportar_favorecidos(conn, indice)  # também injeta 'slug' nos tops
    n_arvore = exportar_arvore(conn, indice)
    execucao = montar_execucao(conn)
    indice["meses"] = exportar_detalhe_mensal(execucao)
    exportar_indices_leves(execucao)
    indice["campos_detalhe"] = CAMPOS_DETALHE
    with open(JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(indice, f, ensure_ascii=False, separators=(",", ":"))

    exportar_csv(execucao)
    try:
        exportar_xlsx(execucao)
        xlsx_msg = os.path.basename(XLSX_PATH)
    except ImportError:
        xlsx_msg = "(openpyxl ausente — XLSX pulado)"
    conn.close()

    print(f"Export concluído: {indice['totais']['pagamentos']} pagamentos (visão caixa), "
          f"{len(execucao)} linhas de execução por empenho, "
          f"R$ {indice['totais']['geral']:,.2f}, {len(indice['alertas'])} alertas, "
          f"{len(indice['meses'])} meses, {n_fav_raiox} raio-X de favorecidos, "
          f"árvore com {n_arvore} funções → "
          f"{os.path.basename(JSON_PATH)}, dados/, favorecidos/, {os.path.basename(CSV_PATH)}, {xlsx_msg}.", file=sys.stderr)


if __name__ == "__main__":
    main()
