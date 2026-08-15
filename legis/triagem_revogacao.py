# -*- coding: utf-8 -*-
"""Triagem do acervo legislativo de Santos: categoriza as normas (eixo temático +
natureza do ato) e sinaliza candidatas a revogação, sob a ótica de um mandato de
oposição comprometido com Estado enxuto, desburocratização e limpeza do acervo.

ATENÇÃO: a Base de Legislação é só de METADADOS (tipo, número, ano, ementa, tags) —
não tem texto integral nem a situação vigente/revogada. Portanto "passível de
revogação" é TRIAGEM por heurística (candidata a verificação), não confirmação.

Entrada : legis-index.json (raiz, UTF-8 limpo)
Saída   : legis/triagem-revogacao.xlsx
"""
import json, re, unicodedata, collections
from pathlib import Path
from datetime import date

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

RAIZ = Path(__file__).resolve().parent.parent
FONTE = RAIZ / "legis-index.json"
SAIDA = Path(__file__).resolve().parent / "triagem-revogacao.xlsx"


def na(s: str) -> str:
    """normaliza: sem acento, minúsculo."""
    return unicodedata.normalize("NFKD", s or "").encode("ascii", "ignore").decode().lower()


def numero_de(item) -> str:
    if (item.get("numero") or "").strip():
        return item["numero"].strip()
    # extrai do título: "Lei n.º 3.850, de ..." / "Decreto n.º 11.161, de ..."
    m = re.search(r"n[.º°o\s]*\s*([\d.]+)", item.get("titulo", ""), re.I)
    return m.group(1).strip(".") if m else ""


# ---------------------------------------------------------------------------
# Classificação. Cada norma recebe (natureza, eixo_tematico, tier, justificativa).
# A ordem das regras importa: a primeira que casar define a NATUREZA do ato.
# ---------------------------------------------------------------------------

def classificar(item):
    blob = na(item.get("tags", "") + " | " + item.get("ementa", "") + " | " + item.get("titulo", ""))
    tipo = na(item.get("tipo", ""))
    tem = lambda *ps: any(p in blob for p in ps)

    # --- NATUREZA do ato + tier de revogação --------------------------------

    # 1) Denominação de logradouro / próprio público
    if tem("denomin", "passa a denominar", "logradouro", "passa a chamar"):
        return ("Denominação de logradouro/próprio",
                "Honorífica / Simbólica", "Baixa",
                "Ato permanente e inócuo; sem custo. Revogar só se houver homenageado controverso.")

    # 2) Data comemorativa / calendário / semana / dia
    if tem("calendario", "datas comemorativas", "comemorativ", "semana ",
           "dia do", "dia da", "dia de", "dia municipal", "dia estadual", "dia nacional",
           "jubileu", "festa ", "festival", "feira ", "campanha ", "mes ", "aniversario") \
            and not tem("posturas", "tributo", "credito"):
        return ("Data comemorativa / Calendário",
                "Honorífica / Simbólica", "Alta",
                "Lei declaratória sem efeito normativo. Candidata a consolidação num "
                "calendário único e revogação do acervo simbólico esparso.")

    # 3) Título honorífico / medalha
    if tem("titulo de cidad", "honra ao merito", "medalha", "comenda"):
        return ("Título / Honraria",
                "Honorífica / Simbólica", "Baixa",
                "Honraria pontual já consumada; sem efeito normativo permanente.")

    # 4) Declaração de utilidade pública
    if tem("utilidade publica", "de utilidade"):
        return ("Declaração de utilidade pública",
                "Terceiro setor", "Alta",
                "Verificar se a entidade ainda existe e atua. UP a entidade extinta/"
                "inativa deve ser revogada (saneamento do acervo).")

    # 5) Repasse / fomento ao terceiro setor
    if tem("termo de fomento", "fomento", "convenio", "subvencao", "subvencional",
           "repasse", "colaboracao", "doacao"):
        return ("Repasse / Fomento (terceiro setor)",
                "Terceiro setor / Fiscal", "Média",
                "Repasse a entidade privada: revisar vigência, contrapartida e prestação "
                "de contas. Autorizações exauridas podem ser revogadas.")

    # 5b) Orçamentário: crédito adicional / suplementação / PPA / LOA / LDO
    if tem("credito adicional", "credito especial", "credito suplementar", "suplementacao",
           "abertura de credito", "plano plurianual", "plurianual", "orcamentaria anual",
           "diretrizes orcamentarias", " ppa", " loa", " ldo"):
        if tem("credito", "suplementacao"):
            return ("Crédito adicional / Suplementação",
                    "Orçamento", "Alta",
                    "Crédito orçamentário de exercício passado, já exaurido. Sem objeto "
                    "atual — candidato a revogação formal no saneamento do acervo.")
        return ("Peça orçamentária (PPA/LDO/LOA)",
                "Orçamento", "Baixa",
                "Peça de planejamento de ciclo vencido; substituída pela peça do ciclo seguinte.")

    # 6) Criação de órgão / conselho / fundo / programa / plano
    if tem("conselho", "fundo ", "fundo municipal", "cria o programa", "institui o programa",
           "programa municipal", "plano municipal", "comite ", "camara tematica",
           "autarquia", "coordenadoria", "ouvidoria"):
        return ("Criação de órgão / programa / fundo",
                "Estrutura administrativa", "Média",
                "Alvo de fiscalização: checar se o órgão/programa funciona, tem custeio e "
                "função clara. Estrutura sem atividade deve ser extinta.")

    # 7) Pessoal / cargos / remuneração
    if tem("cargos", "quadro de pessoal", "gratificacao", "reajuste", "vencimentos",
           "subsidio", "servidores", "carreira", "plano de cargos"):
        return ("Pessoal / Remuneração",
                "Administração / Fiscal", "Baixa",
                "Efeito normativo sobre a folha; revogação isolada rara. Avaliar no conjunto "
                "do estatuto/plano de carreira.")

    # 8) Tributária / fiscal
    if tem("iptu", "iss", "tributo", "tributar", "refis", "debitos", "isencao",
           "remissao", "anistia", "taxa", "contribuicao de melhoria"):
        return ("Tributária / Fiscal",
                "Tributação", "Baixa",
                "Efeito fiscal direto. Benefícios fiscais de prazo vencido (REFIS/anistia) "
                "já exauridos são candidatos a revogação formal.")

    # 9) Urbanismo / posturas / obras
    if tem("posturas", "zoneamento", "uso e ocupacao", "plano diretor", "edificacao",
           "loteamento", "parcelamento", "obras", "codigo de obras"):
        return ("Urbanismo / Posturas / Obras",
                "Urbanismo", "Baixa",
                "Núcleo normativo da cidade; revisar por consolidação, não revogação avulsa.")

    # 10) Alteração / revogação de norma existente
    if tem("altera", "acrescenta", "modifica", "revoga", "da nova redacao", "consolidacao"):
        return ("Alteração / Consolidação legislativa",
                "Técnica legislativa", "Baixa",
                "Ato que altera outra norma; integra-se ao texto-base. Sem revogação própria.")

    # 11) Autorização ao Executivo (genérica)
    if tem("autoriza o poder executivo", "autoriza o municipio", "fica autorizado", "autoriza a"):
        return ("Autorização ao Executivo",
                "Administração", "Média",
                "Autorização pode já ter sido executada/exaurida — verificar e revogar se sem objeto.")

    # 12) Obrigação / regulação sobre o setor privado
    if tem("fica obrigad", "obrigatoriedade", "obrigatorio", "afixar", "placa",
           "estabelecimentos", "cartaz", "cardapio", "comercio", "exibicao obrigatoria",
           "disponibilizar", "informar ao consumidor"):
        return ("Obrigação ao setor privado",
                "Regulação econômica", "Média",
                "Cria obrigação/custo ao setor privado. Revisar sob a ótica da "
                "desburocratização — candidata a revogação se redundante ou de baixa eficácia.")

    # 13) Eixos setoriais (refina o tema; mantém efeito normativo => tier Baixa)
    setoriais = [
        ("saude", "saude"), ("hospital", "saude"), ("vacina", "saude"), ("doenca", "saude"),
        ("educacao", "educacao"), ("escola", "educacao"), ("ensino", "educacao"),
        ("creche", "educacao"), ("aluno", "educacao"),
        ("assistencia", "assistência social"), ("idoso", "assistência social"),
        ("crianca", "assistência social"), ("deficiencia", "assistência social"),
        ("mulher", "assistência social"), ("social", "assistência social"),
        ("animal", "meio ambiente / animais"), ("animais", "meio ambiente / animais"),
        ("ambiental", "meio ambiente / animais"), ("arvore", "meio ambiente / animais"),
        ("residuo", "meio ambiente / animais"),
        ("transporte", "mobilidade"), ("transito", "mobilidade"), ("taxi", "mobilidade"),
        ("estacionamento", "mobilidade"), ("ciclo", "mobilidade"),
        ("seguranca", "segurança"), ("guarda municipal", "segurança"), ("violencia", "segurança"),
        ("cultura", "cultura / esporte"), ("cultural", "cultura / esporte"),
        ("esporte", "cultura / esporte"), ("turismo", "cultura / esporte"),
        ("sepultura", "serviços urbanos"), ("cemiterio", "serviços urbanos"),
        ("habitacao", "habitação"), ("moradia", "habitação"),
    ]
    for chave, eixo_set in setoriais:
        if chave in blob:
            return ("Norma setorial",
                    eixo_set.capitalize(), "Baixa",
                    "Efeito normativo setorial vigente; manter salvo revisão temática específica.")

    return ("Norma setorial / Outra",
            "Setorial / Outra", "Baixa",
            "Sem palavras-chave de categorização; classificar manualmente se relevante.")


def main():
    itens = json.loads(FONTE.read_text(encoding="utf-8"))
    linhas = []
    cont_nat = collections.Counter()
    cont_eixo = collections.Counter()
    cont_tier = collections.Counter()

    for it in itens:
        natureza, eixo, tier, just = classificar(it)
        cont_nat[natureza] += 1
        cont_eixo[eixo] += 1
        cont_tier[tier] += 1
        resumo = (it.get("ementa") or "").strip() or (it.get("tags") or "").strip()
        linhas.append({
            "tipo": it.get("tipo", ""),
            "numero": numero_de(it),
            "ano": it.get("ano", ""),
            "natureza": natureza,
            "eixo": eixo,
            "tier": tier,
            "just": just,
            "resumo": resumo[:300],
            "url": it.get("url_legis", ""),
            "pdf": it.get("url_pdf", ""),
        })

    # ordena: tier (Alta>Média>Baixa), depois ano desc, depois tipo/numero
    ordem_tier = {"Alta": 0, "Média": 1, "Baixa": 2}
    linhas.sort(key=lambda l: (ordem_tier[l["tier"]], -(l["ano"] or 0), l["tipo"], l["numero"]))

    gerar_xlsx(linhas, cont_nat, cont_eixo, cont_tier, len(itens))
    print(f"OK -> {SAIDA}  ({len(itens)} normas)")
    print("Por tier:", dict(cont_tier))
    print("Por natureza:")
    for k, v in cont_nat.most_common():
        print(f"  {v:5d}  {k}")


# ---------------------------------------------------------------------------
# Geração do XLSX (navy/dourado, alinhado à identidade do projeto)
# ---------------------------------------------------------------------------
NAVY = "1B2A4A"
GOLD = "C9A24B"
CINZA = "F2F2F2"
TIER_FILL = {"Alta": "F4B6B6", "Média": "FBE3A1", "Baixa": "D7E4C9"}
borda = Border(*[Side(style="thin", color="DDDDDD")] * 4)


def _cabecalho(ws, headers, larguras):
    f_h = Font(bold=True, color="FFFFFF", size=10)
    fill = PatternFill("solid", fgColor=NAVY)
    for c, (h, w) in enumerate(zip(headers, larguras), 1):
        cell = ws.cell(1, c, h)
        cell.font = f_h
        cell.fill = fill
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"


def gerar_xlsx(linhas, cont_nat, cont_eixo, cont_tier, total):
    wb = openpyxl.Workbook()

    # ---- Aba 1: Resumo --------------------------------------------------
    ws = wb.active
    ws.title = "Resumo"
    ws["A1"] = "Triagem do Acervo Legislativo — Câmara de Santos"
    ws["A1"].font = Font(bold=True, size=14, color=NAVY)
    ws["A2"] = f"Gerado em {date.today():%d/%m/%Y} · {total} normas (Base de Legislação / Legis Prefeitura)"
    ws["A2"].font = Font(italic=True, size=9, color="666666")
    ws["A4"] = ("ATENÇÃO — METODOLOGIA: a base contém só metadados (tipo, número, ano, ementa, tags), "
                "sem texto integral nem situação vigente/revogada. A coluna \"Passível de revogação\" é "
                "TRIAGEM por heurística (candidata a verificação manual), não confirmação jurídica. "
                "Alta = forte indício de obsolescência / acervo simbólico; Média = revisar necessidade e "
                "vigência; Baixa = efeito normativo, manter salvo revisão temática.")
    ws["A4"].font = Font(size=9, color="993333")
    ws["A4"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A4:E7")

    r = 9
    ws.cell(r, 1, "Por grau de revogação").font = Font(bold=True, color=NAVY, size=11)
    r += 1
    for k in ("Alta", "Média", "Baixa"):
        ws.cell(r, 1, k).fill = PatternFill("solid", fgColor=TIER_FILL[k])
        ws.cell(r, 1).font = Font(bold=True)
        ws.cell(r, 2, cont_tier.get(k, 0))
        r += 1

    r += 1
    ws.cell(r, 1, "Por natureza do ato").font = Font(bold=True, color=NAVY, size=11)
    r += 1
    for k, v in cont_nat.most_common():
        ws.cell(r, 1, k)
        ws.cell(r, 2, v)
        r += 1

    r2 = 10
    ws.cell(9, 4, "Por eixo temático").font = Font(bold=True, color=NAVY, size=11)
    for k, v in cont_eixo.most_common():
        ws.cell(r2, 4, k)
        ws.cell(r2, 5, v)
        r2 += 1

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 10

    # ---- Aba 2: Triagem completa ---------------------------------------
    ws2 = wb.create_sheet("Triagem completa")
    headers = ["Tipo", "Número", "Ano", "Natureza do ato", "Eixo temático",
               "Passível de revogação", "Justificativa / Ação sugerida",
               "Ementa / Tags", "Link Legis", "PDF"]
    larguras = [16, 9, 6, 30, 22, 14, 50, 60, 24, 10]
    _cabecalho(ws2, headers, larguras)
    f = Font(size=9)
    for i, l in enumerate(linhas, 2):
        vals = [l["tipo"], l["numero"], l["ano"], l["natureza"], l["eixo"],
                l["tier"], l["just"], l["resumo"], l["url"], l["pdf"]]
        for c, v in enumerate(vals, 1):
            cell = ws2.cell(i, c, v)
            cell.font = f
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = borda
        ws2.cell(i, 6).fill = PatternFill("solid", fgColor=TIER_FILL[l["tier"]])
        ws2.cell(i, 6).font = Font(size=9, bold=True)
        if l["url"]:
            ws2.cell(i, 9).hyperlink = l["url"]
            ws2.cell(i, 9).value = "abrir"
            ws2.cell(i, 9).font = Font(size=9, color="0563C1", underline="single")
        if l["pdf"]:
            ws2.cell(i, 10).hyperlink = l["pdf"]
            ws2.cell(i, 10).value = "PDF"
            ws2.cell(i, 10).font = Font(size=9, color="0563C1", underline="single")
    ws2.auto_filter.ref = f"A1:J{len(linhas)+1}"

    # ---- Aba 3: Prioridade (só Alta) -----------------------------------
    altas = [l for l in linhas if l["tier"] == "Alta"]
    ws3 = wb.create_sheet("Prioridade (revisar)")
    _cabecalho(ws3, headers, larguras)
    for i, l in enumerate(altas, 2):
        vals = [l["tipo"], l["numero"], l["ano"], l["natureza"], l["eixo"],
                l["tier"], l["just"], l["resumo"], l["url"], l["pdf"]]
        for c, v in enumerate(vals, 1):
            cell = ws3.cell(i, c, v)
            cell.font = f
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = borda
        ws3.cell(i, 6).fill = PatternFill("solid", fgColor=TIER_FILL["Alta"])
        if l["url"]:
            ws3.cell(i, 9).hyperlink = l["url"]; ws3.cell(i, 9).value = "abrir"
            ws3.cell(i, 9).font = Font(size=9, color="0563C1", underline="single")
        if l["pdf"]:
            ws3.cell(i, 10).hyperlink = l["pdf"]; ws3.cell(i, 10).value = "PDF"
            ws3.cell(i, 10).font = Font(size=9, color="0563C1", underline="single")
    ws3.auto_filter.ref = f"A1:J{len(altas)+1}"

    wb.save(SAIDA)


if __name__ == "__main__":
    main()
