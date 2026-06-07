#!/usr/bin/env python3
"""
Export da base Legis → XLSX, CSV e JSON
=======================================

Lê `legis/legis.sqlite` e gera:
  - legis/legis.xlsx        (planilha p/ gabinete)
  - legis/legis.csv         (análise)
  - legis-index.json (raiz) (índice consumido pelo painel legis.html)

Uso:
    python legis/export.py
"""

import csv
import json
import os
import sqlite3
import sys

for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8")
    except (AttributeError, ValueError):
        pass

AQUI = os.path.dirname(os.path.abspath(__file__))
RAIZ = os.path.dirname(AQUI)
DB_PATH = os.path.join(AQUI, "legis.sqlite")
XLSX_PATH = os.path.join(AQUI, "legis.xlsx")
CSV_PATH = os.path.join(AQUI, "legis.csv")
JSON_PATH = os.path.join(RAIZ, "legis-index.json")

COLUNAS = ["id", "tipo", "numero", "ano", "data_norma", "titulo", "ementa",
           "tags", "situacao", "url_legis", "url_pdf", "fonte", "data_coleta"]


def ler_normas() -> list[dict]:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    rows = conn.execute(
        f"SELECT {','.join(COLUNAS)} FROM normas ORDER BY ano DESC, tipo, numero"
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def exportar_csv(normas: list[dict]) -> None:
    with open(CSV_PATH, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=COLUNAS)
        w.writeheader()
        w.writerows(normas)


def exportar_xlsx(normas: list[dict]) -> None:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Normas"
    ws.append(COLUNAS)
    for n in normas:
        ws.append([n.get(c, "") for c in COLUNAS])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUNAS))}{len(normas)+1}"
    larguras = {"titulo": 50, "ementa": 70, "tags": 30, "url_legis": 45, "url_pdf": 45}
    for i, c in enumerate(COLUNAS, 1):
        ws.column_dimensions[get_column_letter(i)].width = larguras.get(c, 14)
    wb.save(XLSX_PATH)


def exportar_json(normas: list[dict]) -> None:
    # campos enxutos para o painel (sem fonte/data_coleta)
    campos = ["id", "tipo", "numero", "ano", "data_norma", "titulo", "ementa",
              "tags", "situacao", "url_legis", "url_pdf"]
    dados = [{c: n.get(c, "") for c in campos} for n in normas]
    with open(JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(dados, f, ensure_ascii=False, separators=(",", ":"))


def main():
    if not os.path.exists(DB_PATH):
        print(f"Banco não encontrado: {DB_PATH}. Rode o crawler primeiro.", file=sys.stderr)
        sys.exit(1)
    normas = ler_normas()
    exportar_csv(normas)
    exportar_xlsx(normas)
    exportar_json(normas)
    print(f"Export concluído: {len(normas)} normas → "
          f"{os.path.basename(XLSX_PATH)}, {os.path.basename(CSV_PATH)}, "
          f"{os.path.basename(JSON_PATH)}.", file=sys.stderr)


if __name__ == "__main__":
    main()
