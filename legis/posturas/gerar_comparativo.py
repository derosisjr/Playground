# -*- coding: utf-8 -*-
"""
Gera comparativo_posturas.xlsx — comparativo entre o Código de Posturas vigente
(Lei nº 3.531/1968, texto consolidado) e o PLC do Executivo (PA 59314/2025-78).

Fonte de dados: CSVs em legis/posturas/dados/ (editáveis; o script apenas estiliza).
Identidade visual reaproveitada de legis/triagem_revogacao.py.
"""
import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE = Path(__file__).parent
DADOS = BASE / "dados"
SAIDA = BASE / "comparativo_posturas.xlsx"

# ---- identidade visual da casa (cf. legis/triagem_revogacao.py) ----
NAVY = "1B2A4A"
GOLD = "C9A24B"
CINZA = "F2F2F2"
FILL_STATUS = {
    "MANTIDO": "D7E4C9",                      # verde
    "ALTERADO": "FBE3A1",                     # âmbar
    "CONDENSADO": "DCE6F1",                   # azul claro
    "REMETIDO": "EAD9B8",                     # dourado claro
    "REVOGADO_SEM_CORRESPONDENTE": "F4B6B6",  # vermelho
    "JA_REVOGADO": "D9D9D9",                  # cinza
    "NOVO": "C9DFC0",                         # verde claro
}
ROTULO_STATUS = {
    "MANTIDO": "Mantido/equivalente",
    "ALTERADO": "Alterado (mérito)",
    "CONDENSADO": "Condensado",
    "REMETIDO": "Remetido a outra norma",
    "REVOGADO_SEM_CORRESPONDENTE": "Revogado sem correspondente",
    "JA_REVOGADO": "Já revogado anteriormente",
    "NOVO": "Novo no PLC",
}

THIN = Side(style="thin", color="BFBFBF")
BORDA = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# correções de vírgulas escapadas como ';' nos CSVs
_FIXES = {"15.067;25": "15.067,25", "1;50 m": "1,50 m", "2;20 m": "2,20 m"}


def _fix(txt: str) -> str:
    for a, b in _FIXES.items():
        txt = txt.replace(a, b)
    return txt


def _ler(nome):
    with open(DADOS / nome, encoding="utf-8-sig") as f:
        return [{k: _fix(v or "") for k, v in row.items()} for row in csv.DictReader(f)]


def _cabecalho(ws, colunas, larguras):
    for i, (titulo, larg) in enumerate(zip(colunas, larguras), start=1):
        c = ws.cell(row=1, column=i, value=titulo)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDA
        ws.column_dimensions[get_column_letter(i)].width = larg
    ws.freeze_panes = "A2"


def _linha(ws, r, valores, fill=None, wrap=True):
    for i, v in enumerate(valores, start=1):
        c = ws.cell(row=r, column=i, value=v)
        c.font = Font(size=10)
        c.alignment = Alignment(vertical="top", wrap_text=wrap)
        c.border = BORDA
        if fill:
            c.fill = PatternFill("solid", fgColor=fill)


def _titulo_bloco(ws, r, texto, ncols):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    c = ws.cell(row=r, column=1, value=texto)
    c.font = Font(bold=True, color=NAVY, size=11)
    c.fill = PatternFill("solid", fgColor=CINZA)


# ------------------------------------------------------------------ dados
blocos = _ler("blocos_correspondencia.csv")
novos = _ler("novos_plc.csv")
mudancas = _ler("mudancas_substantivas.csv")
revog = _ler("revogacoes.csv")

# expande blocos em linhas artigo-a-artigo
artigos = []
for b in blocos:
    ini, fim = int(b["art_ini"]), int(b["art_fim"])
    for n in range(ini, fim + 1):
        artigos.append({**b, "artigo": n})

wb = Workbook()

# ------------------------------------------------------------ 1. Resumo
ws = wb.active
ws.title = "Resumo"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
for col, w in zip("BCDE", (46, 16, 16, 34)):
    ws.column_dimensions[col].width = w

r = 2
c = ws.cell(row=r, column=2, value="CÓDIGO DE POSTURAS DE SANTOS — COMPARATIVO")
c.font = Font(bold=True, size=15, color=NAVY)
r += 1
c = ws.cell(row=r, column=2, value="Lei nº 3.531/1968 (consolidada)  ×  PLC do Executivo (PA 59314/2025-78)")
c.font = Font(size=11, color=GOLD, bold=True)
r += 2

ressalva = (
    "RESSALVA METODOLÓGICA — Triagem analítica de apoio parlamentar, elaborada por leitura "
    "comparada dos dois textos; não constitui parecer jurídico. Correspondências incertas estão "
    "marcadas [Verificar]. Artigos com sufixo (ex.: 119-A, 249-B, 300-B, 627-A) foram anotados na "
    "linha do artigo-base. Dispositivos já revogados por leis anteriores ao PLC recebem status "
    "próprio e NÃO contam como cortes do PLC."
)
ws.merge_cells(start_row=r, start_column=2, end_row=r + 3, end_column=5)
c = ws.cell(row=r, column=2, value=ressalva)
c.font = Font(size=10, italic=True)
c.alignment = Alignment(wrap_text=True, vertical="top")
c.fill = PatternFill("solid", fgColor=CINZA)
r += 5

# contagens por status
cont = {}
for a in artigos:
    cont[a["status"]] = cont.get(a["status"], 0) + 1

c = ws.cell(row=r, column=2, value="Artigos do código vigente por destino no PLC")
c.font = Font(bold=True, size=12, color=NAVY)
r += 1
for st in ("MANTIDO", "ALTERADO", "CONDENSADO", "REMETIDO",
           "REVOGADO_SEM_CORRESPONDENTE", "JA_REVOGADO"):
    ws.cell(row=r, column=2, value=ROTULO_STATUS[st]).font = Font(size=10)
    q = ws.cell(row=r, column=3, value=cont.get(st, 0))
    q.font = Font(bold=True, size=10)
    q.fill = PatternFill("solid", fgColor=FILL_STATUS[st])
    q.alignment = Alignment(horizontal="center")
    r += 1
ws.cell(row=r, column=2, value="Total de artigos mapeados (1–640)").font = Font(bold=True, size=10)
q = ws.cell(row=r, column=3, value=sum(cont.values()))
q.font = Font(bold=True, size=10)
r += 2

for rotulo, valor in (
    ("Artigos do PLC", "335"),
    ("Blocos temáticos novos no PLC", str(len(novos))),
    ("Mudanças substantivas catalogadas", str(len(mudancas))),
    ("Normas revogadas nominalmente (art. 334 do PLC)", str(len(revog))),
):
    ws.cell(row=r, column=2, value=rotulo).font = Font(size=10)
    q = ws.cell(row=r, column=3, value=valor)
    q.font = Font(bold=True, size=10)
    q.alignment = Alignment(horizontal="center")
    r += 1
r += 1

c = ws.cell(row=r, column=2, value="Legenda de status")
c.font = Font(bold=True, size=12, color=NAVY)
r += 1
for st, rot in ROTULO_STATUS.items():
    q = ws.cell(row=r, column=2, value=rot)
    q.font = Font(size=10)
    q.fill = PatternFill("solid", fgColor=FILL_STATUS[st])
    r += 1

# ------------------------------------------- 2. Mapa por Título/Capítulo
ws = wb.create_sheet("Mapa por Título")
COLS = ("Estrutura do código vigente (Lei 3.531/68)", "Artigos",
        "Destino no PLC", "Artigos PLC", "Síntese")
_cabecalho(ws, COLS, (38, 12, 34, 12, 60))
MAPA = [
    ("Tít. I — Disposições Gerais", "1–4", "Tít. I — Disposições Gerais (ampliado)", "1–10",
     "Ganha objetivos, princípios, competências e direitos/deveres dos munícipes."),
    ("Tít. II — Higiene Pública", "5–185", "Tít. II — Higiene Pública", "11–104",
     "Fortemente condensado; ganha zoonoses, coleta seletiva e poluição ampliada; cemitérios ficam de fora (matéria já revogada em 2011)."),
    ("— (sem correspondente)", "—", "Tít. III — Drenagem Urbana (NOVO)", "105–127",
     "Título inteiramente novo: obrigações de imóveis, construtoras e concessionárias."),
    ("Tít. III — Bem-Estar Público", "186–308", "Tít. IV — Bem-Estar Público", "128–220",
     "Ganha capítulos de parques, calçadas e mobiliário urbano; publicidade e sossego condensados; animais atualizado."),
    ("Tít. IV — Instalações Elétricas e Mecânicas", "309–426", "Removido (elevadores: Tít. IV Cap. XIII)", "217–220",
     "118 artigos suprimidos; remissão a ABNT, Corpo de Bombeiros e Código de Edificações."),
    ("Tít. V — Localização e Funcionamento", "427–572", "Tít. V — Localização e Funcionamento", "221–279",
     "Horário desregulamentado; ambulantes e bancas em leis específicas; inflamáveis, segurança do trabalho e pesos/medidas removidos; ganha capítulo de concessionárias."),
    ("Tít. VI — Fiscalização da Prefeitura", "573–585", "Tít. VI — Fiscalização", "280–298",
     "Procedimentos modernizados; ganha termo de compromisso (293–298)."),
    ("Tít. VII — Aceitação das Instalações", "586–588", "Removido", "—",
     "Etapa absorvida pelo licenciamento."),
    ("Tít. VIII — Infrações e Penalidades", "589–627-B", "Tít. VII — Infrações e Penalidades", "299–330",
     "Multas em UFESP com dosimetria; advertência, apreensão, suspensão, embargo, interdição, lacração, emparedamento, cassação e demolição sistematizados."),
    ("Tít. IX — Disposições Finais", "628–640", "Tít. VIII — Disposições Finais", "331–335",
     "Comissão Consultiva do Código desaparece; cláusula revogatória nominal (art. 334)."),
]
for i, row in enumerate(MAPA, start=2):
    _linha(ws, i, row, fill=CINZA if i % 2 == 0 else None)
ws.auto_filter.ref = f"A1:E{len(MAPA) + 1}"

# --------------------------------------------- 3. Mudanças Substantivas
ws = wb.create_sheet("Mudanças Substantivas")
COLS = ("Tema", "Código atual", "PLC", "Tipo", "Síntese", "Ponto de atenção")
_cabecalho(ws, COLS, (28, 14, 14, 16, 70, 12))
for i, m in enumerate(mudancas, start=2):
    fill = "FBE3A1" if m["atencao"].strip().upper() == "SIM" else None
    _linha(ws, i, (m["tema"], m["art_atual"], m["art_plc"], m["tipo"], m["sintese"],
                   m["atencao"].strip().upper()), fill=fill)
ws.auto_filter.ref = f"A1:F{len(mudancas) + 1}"

# --------------------------------------- 4. Correspondência Artigo-a-Artigo
ws = wb.create_sheet("Artigo a Artigo")
COLS = ("Art. atual", "Localização no código vigente", "Tema", "Status",
        "Art(s). no PLC", "Observação")
_cabecalho(ws, COLS, (10, 26, 34, 26, 14, 64))
for i, a in enumerate(artigos, start=2):
    _linha(ws, i, (a["artigo"], a["capitulo_atual"], a["tema"],
                   ROTULO_STATUS[a["status"]], a["plc_arts"], a["observacao"]),
           fill=FILL_STATUS[a["status"]])
ws.auto_filter.ref = f"A1:F{len(artigos) + 1}"

# ----------------------------------------------------- 5. Novos no PLC
ws = wb.create_sheet("Novos no PLC")
COLS = ("Art(s). PLC", "Tema", "Síntese", "Ponto de atenção")
_cabecalho(ws, COLS, (12, 34, 80, 12))
for i, n in enumerate(novos, start=2):
    fill = "FBE3A1" if n["atencao"].strip().upper() == "SIM" else FILL_STATUS["NOVO"]
    _linha(ws, i, (n["plc_arts"], n["tema"], n["sintese"], n["atencao"].strip().upper()),
           fill=fill)
ws.auto_filter.ref = f"A1:D{len(novos) + 1}"

# ------------------------------------------------- 6. Revogações (art. 334)
ws = wb.create_sheet("Revogações Expressas")
COLS = ("Norma", "Ano", "Observação")
_cabecalho(ws, COLS, (34, 10, 60))
for i, rv in enumerate(revog, start=2):
    _linha(ws, i, (rv["norma"], rv["ano"], rv["observacao"]),
           fill=CINZA if i % 2 == 0 else None)
ws.auto_filter.ref = f"A1:C{len(revog) + 1}"

# também grava a correspondência expandida em CSV (auditoria)
with open(DADOS / "correspondencia.csv", "w", newline="", encoding="utf-8-sig") as f:
    w = csv.writer(f)
    w.writerow(["artigo", "capitulo_atual", "tema", "status", "plc_arts", "observacao"])
    for a in artigos:
        w.writerow([a["artigo"], a["capitulo_atual"], a["tema"],
                    a["status"], a["plc_arts"], a["observacao"]])

wb.save(SAIDA)
print(f"OK — {SAIDA.name}")
print(f"Artigos mapeados: {sum(cont.values())} (esperado: 640)")
for st, q in sorted(cont.items(), key=lambda x: -x[1]):
    print(f"  {ROTULO_STATUS[st]:<32} {q}")
print(f"Mudanças substantivas: {len(mudancas)} | Novos: {len(novos)} | Revogações: {len(revog)}")
