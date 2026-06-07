#!/usr/bin/env python3
"""
Export da base de Proposituras → XLSX, CSV e JSON
=================================================

Lê `proposituras/proposituras.sqlite` e gera:
  - proposituras/proposituras.xlsx   (planilha p/ gabinete; abas Proposituras + Tramitação)
  - proposituras/proposituras.csv    (análise)
  - proposituras-index.json (raiz)   (índice consumido pelo painel proposituras.html)

Uso:
    python proposituras/export.py
"""

import csv
import json
import os
import sqlite3
import sys

for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8")
    except (AttributeError, ValueError):
        pass

AQUI = os.path.dirname(os.path.abspath(__file__))
RAIZ = os.path.dirname(AQUI)
DB_PATH = os.path.join(AQUI, "proposituras.sqlite")
XLSX_PATH = os.path.join(AQUI, "proposituras.xlsx")
CSV_PATH = os.path.join(AQUI, "proposituras.csv")
JSON_PATH = os.path.join(RAIZ, "proposituras-index.json")

COLUNAS = ["cod", "tipo", "subtipo", "numero", "ano", "processo", "data_propositura",
           "autor", "ementa", "local_atual", "situacao", "data_situacao",
           "url_detalhes", "url_pdf", "fonte", "data_coleta"]


def ler_proposituras() -> list[dict]:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    rows = conn.execute(
        f"SELECT {','.join(COLUNAS)} FROM proposituras ORDER BY ano DESC, subtipo, numero"
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def ler_tramitacao() -> list[dict]:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    rows = conn.execute(
        "SELECT prop_cod, ordem, data, local, despacho, url_arquivo "
        "FROM tramitacao ORDER BY prop_cod, ordem"
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def exportar_csv(props: list[dict]) -> None:
    with open(CSV_PATH, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=COLUNAS)
        w.writeheader()
        w.writerows(props)


def exportar_xlsx(props: list[dict], tramites: list[dict]) -> None:
    from openpyxl import Workbook
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
    from openpyxl.utils import get_column_letter

    def limpar(v):
        # remove caracteres de controle que o openpyxl recusa (ex.: em despachos)
        return ILLEGAL_CHARACTERS_RE.sub("", v) if isinstance(v, str) else v

    wb = Workbook()
    ws = wb.active
    ws.title = "Proposituras"
    ws.append(COLUNAS)
    for p in props:
        ws.append([limpar(p.get(c, "")) for c in COLUNAS])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUNAS))}{len(props)+1}"
    larguras = {"subtipo": 26, "ementa": 70, "autor": 28, "local_atual": 20,
                "situacao": 18, "url_detalhes": 45, "url_pdf": 45}
    for i, c in enumerate(COLUNAS, 1):
        ws.column_dimensions[get_column_letter(i)].width = larguras.get(c, 14)

    ws2 = wb.create_sheet("Tramitação")
    cols2 = ["prop_cod", "ordem", "data", "local", "despacho", "url_arquivo"]
    ws2.append(cols2)
    for t in tramites:
        ws2.append([limpar(t.get(c, "")) for c in cols2])
    ws2.freeze_panes = "A2"
    for i, c in enumerate(cols2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = {"despacho": 80, "url_arquivo": 45}.get(c, 16)

    wb.save(XLSX_PATH)


def exportar_json(props: list[dict]) -> None:
    # campos enxutos para o painel (sem fonte/data_coleta/processo)
    campos = ["cod", "tipo", "subtipo", "numero", "ano", "data_propositura",
              "autor", "ementa", "local_atual", "situacao", "url_detalhes", "url_pdf"]
    dados = [{c: p.get(c, "") for c in campos} for p in props]
    with open(JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(dados, f, ensure_ascii=False, separators=(",", ":"))


def main():
    if not os.path.exists(DB_PATH):
        print(f"Banco não encontrado: {DB_PATH}. Rode o crawler primeiro.", file=sys.stderr)
        sys.exit(1)
    props = ler_proposituras()
    tramites = ler_tramitacao()
    exportar_csv(props)
    exportar_xlsx(props, tramites)
    exportar_json(props)
    print(f"Export concluído: {len(props)} proposituras ({len(tramites)} despachos) → "
          f"{os.path.basename(XLSX_PATH)}, {os.path.basename(CSV_PATH)}, "
          f"{os.path.basename(JSON_PATH)}.", file=sys.stderr)


if __name__ == "__main__":
    main()
